import ast
import io
import os
from datetime import datetime
from sqlalchemy.orm import Session
from sqlalchemy.dialects.postgresql import insert as pg_insert
import sqlalchemy as sa
from app.models.tecdoc import SupplierPrice
from app.models.imports import PriceImport
from app.models.parts import Part
from app.models.suppliers import Supplier, SupplierOffer
from app.core.db import TecDocSessionLocal


IMPORT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "imports")


def _ensure_import_dir():
    os.makedirs(IMPORT_DIR, exist_ok=True)


def build_xlsx_from_json(items: list) -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Prices"
    headers = ["cid", "article", "brand", "category", "name", "price", "currency",
               "stock_total", "stock_regions", "tecdoc_article", "image_url"]
    if not items:
        ws.append(headers)
    else:
        ws.append(headers)
        for item in items:
            # Try multiple possible field names for image
            img = ""
            for img_key in ("image_path", "image", "photo", "picture", "image_url", "img", "product_image", "photo_url", "зображення_товару"):
                val = item.get(img_key, "")
                if val:
                    img = val
                    break
            row = [
                item.get("cid", ""),
                item.get("article", ""),
                item.get("brand", ""),
                item.get("category", ""),
                item.get("name", ""),
                item.get("price_type_10", item.get("price_currency_980", "")),
                "UAH",
                _sum_stock(item),
                str(_extract_stock(item)),
                item.get("tecdoc_article", ""),
                img,
            ]
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _sum_stock(item: dict) -> int:
    total = 0
    for key, val in item.items():
        if key.startswith("count_warehouse_"):
            try:
                v = int(float(str(val).replace(">", "").replace("<", "").strip()))
                total += v
            except (ValueError, TypeError):
                pass
    return total


def _extract_stock(item: dict) -> dict:
    regions = {}
    for key, val in item.items():
        if key.startswith("count_warehouse_"):
            try:
                v = int(float(str(val).replace(">", "").replace("<", "").strip()))
            except (ValueError, TypeError):
                v = 0
            warehouse = key.replace("count_warehouse_", "")
            regions[warehouse] = v
    return regions


def _resolve_tecdoc_brand(tecdoc_db, brand_name: str) -> int | None:
    if not brand_name or not tecdoc_db:
        return None
    from sqlalchemy import text
    norm = brand_name.lower().strip()
    row = tecdoc_db.execute(
        text("SELECT id FROM autodb_suppliers WHERE LOWER(normalized_name) = :name OR LOWER(matchcode) = :match LIMIT 1"),
        {"name": norm, "match": norm},
    ).first()
    if row:
        return row[0]
    row = tecdoc_db.execute(
        text("SELECT id FROM autodb_suppliers WHERE LOWER(name) = :name LIMIT 1"),
        {"name": norm},
    ).first()
    return row[0] if row else None


def parse_xlsx_to_prices(db: Session, supplier: str, file_data: bytes, tecdoc_db: Session | None = None) -> int:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_data), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0

    headers = [str(h).strip().lower() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    brand_cache = {}
    batch = []
    total = 0

    def flush_batch():
        nonlocal batch
        if not batch:
            return
        stmt = pg_insert(SupplierPrice).values(batch)
        stmt = stmt.on_conflict_do_update(
            index_elements=["supplier", "article"],
            set_={
                "brand": stmt.excluded.brand,
                "name": stmt.excluded.name,
                "price": stmt.excluded.price,
                "currency": stmt.excluded.currency,
                "stock_total": stmt.excluded.stock_total,
                "stock_regions": stmt.excluded.stock_regions,
                "tecdoc_article": stmt.excluded.tecdoc_article,
                "tecdoc_brand_id": stmt.excluded.tecdoc_brand_id,
                "match_status": stmt.excluded.match_status,
                "category": stmt.excluded.category,
                "image_url": stmt.excluded.image_url,
            },
        )
        db.execute(stmt)
        db.commit()
        batch = []

    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        values = {headers[i]: row[i] if row[i] is not None else "" for i in range(min(len(headers), len(row)))}

        article = str(values.get("article", values.get("articul", values.get("oem", ""))))
        if not article or article.lower() == "article":
            continue

        brand = str(values.get("brand", values.get("producer", values.get("category", ""))))
        name = str(values.get("name", values.get("title", values.get("description", ""))))
        price_str = str(values.get("price", values.get("price_currency_980", values.get("ррц", "0")))).replace(",", ".")
        try:
            price = float(price_str) if price_str else None
        except (ValueError, TypeError):
            price = None

        stock_total = 0
        stock_regions = {}

        sr_val = values.get("stock_regions")
        if sr_val and isinstance(sr_val, str) and sr_val.strip().startswith("{"):
            try:
                parsed = ast.literal_eval(sr_val)
                if isinstance(parsed, dict):
                    stock_regions.update(parsed)
            except (ValueError, SyntaxError):
                pass

        st_val = values.get("stock_total")
        if st_val is not None:
            try:
                stock_total = int(float(str(st_val).replace(">", "").replace("<", "").strip()))
            except (ValueError, TypeError):
                pass

        for k, v in values.items():
            if k in ("stock_total", "stock_regions"):
                continue
            if any(kw in k for kw in ["count", "remain", "warehouse", "склад"]):
                try:
                    qty = int(float(str(v).replace(">", "").replace("<", "").strip()))
                    if k not in stock_regions:
                        stock_regions[k] = qty
                except (ValueError, TypeError):
                    pass

        currency = str(values.get("currency", values.get("currency_code", "UAH"))).upper()

        tecdoc_article = str(values.get("tecdoc_article", "")).strip() or None

        if brand not in brand_cache:
            brand_cache[brand] = _resolve_tecdoc_brand(tecdoc_db, brand) if tecdoc_db else None
        tecdoc_brand_id = brand_cache.get(brand)
        match_status = "matched" if tecdoc_brand_id else "pending"

        category = str(values.get("category", "")).strip() or None
        image_url = str(values.get("image_url", "")).strip() or None

        batch.append({
            "supplier": supplier,
            "article": str(article)[:100],
            "brand": str(brand)[:100] if brand else None,
            "name": str(name)[:500] if name else None,
            "price": price,
            "currency": currency,
            "stock_total": stock_total,
            "stock_regions": stock_regions if stock_regions else None,
            "tecdoc_article": tecdoc_article,
            "tecdoc_brand_id": tecdoc_brand_id,
            "match_status": match_status,
            "category": category,
            "image_url": image_url,
        })
        total += 1

        if len(batch) >= 1000:
            flush_batch()

    flush_batch()
    return total


def promote_all_to_catalog(db: Session, supplier: str, progress_cb=None):
    supplier_obj = db.query(Supplier).filter(Supplier.name == supplier).first()
    if not supplier_obj:
        return 0

    rows = db.query(SupplierPrice).filter(
        SupplierPrice.supplier == supplier,
        SupplierPrice.price.isnot(None),
    ).all()

    if not rows:
        return 0

    total = len(rows)
    part_batch = []
    part_ids = {}

    # First pass: try to match by tecdoc_article (cross-supplier dedup)
    for sp in rows:
        brand = sp.brand or ""
        key = (sp.article, brand)
        if key not in part_ids:
            existing = None
            if sp.tecdoc_article:
                existing = db.query(Part).filter(
                    Part.article == sp.tecdoc_article,
                    Part.brand == brand,
                ).first()
            if existing:
                existing.supplier_article = sp.article
                if sp.tecdoc_brand_id:
                    existing.brand_id = sp.tecdoc_brand_id
                part_ids[key] = existing.id
            else:
                part_batch.append({
                    "article": sp.tecdoc_article or sp.article,
                    "supplier_article": sp.article,
                    "brand": brand,
                    "name": sp.name or sp.article,
                    "brand_id": sp.tecdoc_brand_id or 0,
                    "sku": sp.sku,
                    "is_active": True,
                    "image_url": sp.image_url,
                })
                part_ids[key] = None

	    for batch_start in range(0, len(part_batch), 1000):
        chunk = part_batch[batch_start:batch_start + 1000]
        stmt = pg_insert(Part).values(chunk)
        stmt = stmt.on_conflict_do_update(
            index_elements=["article", sa.text("COALESCE(brand, '')")],
            set_={
                "name": stmt.excluded.name,
                "brand_id": stmt.excluded.brand_id,
                "sku": stmt.excluded.sku,
                "is_active": True,
                "image_url": stmt.excluded.image_url,
            },
        )
        db.execute(stmt)
        db.commit()
        if progress_cb:
            progress_cb(10 + int(30 * batch_start / max(len(part_batch), 1)))
    db.commit()  # commit any existing Part updates from cross-supplier matching

    for part in db.query(Part).all():
        key = (part.supplier_article or part.article, part.brand or "")
        if key in part_ids:
            part_ids[key] = part.id

    offer_batch = []
    for sp in rows:
        key = (sp.article, sp.brand or "")
        part_id = part_ids.get(key)
        if not part_id:
            continue
        offer_batch.append({
            "part_id": part_id,
            "supplier_id": supplier_obj.id,
            "price": sp.price,
            "currency": sp.currency or "UAH",
            "quantity": sp.stock_total,
            "stock_regions": sp.stock_regions,
            "updated_at": datetime.utcnow(),
        })

    matched = 0
    for batch_start in range(0, len(offer_batch), 1000):
        chunk = offer_batch[batch_start:batch_start + 1000]
        stmt = pg_insert(SupplierOffer).values(chunk)
        stmt = stmt.on_conflict_do_update(
            index_elements=["part_id", "supplier_id"],
            set_={
                "price": stmt.excluded.price,
                "currency": stmt.excluded.currency,
                "quantity": stmt.excluded.quantity,
                "stock_regions": stmt.excluded.stock_regions,
                "updated_at": stmt.excluded.updated_at,
            },
        )
        db.execute(stmt)
        matched += len(chunk)
        if progress_cb:
            progress_cb(40 + int(60 * batch_start / max(len(offer_batch), 1)))

    db.commit()
    return matched


def assign_gpl_categories(db: Session, supplier: str):
    from app.services.gpl_categories import GPL_CATEGORY_MAP
    from sqlalchemy import text

    if supplier.upper() != "GPL":
        return

    rows = db.query(SupplierPrice).filter(
        SupplierPrice.supplier == supplier,
        SupplierPrice.category.isnot(None),
        SupplierPrice.category != "",
    ).all()

    cat_map = {}
    for sp in rows:
        tecdoc_id = GPL_CATEGORY_MAP.get(sp.category)
        if tecdoc_id is None:
            continue
        cat_map.setdefault(tecdoc_id, []).append((sp.article, sp.brand or "", sp.name or ""))

    if not cat_map:
        return 0

    all_tecdoc_ids = list(cat_map.keys())
    tecdoc_to_cat = {}
    for i in range(0, len(all_tecdoc_ids), 500):
        chunk = all_tecdoc_ids[i:i + 500]
        cat_rows = db.execute(
            text(f"SELECT id, tecdoc_id FROM part_categories WHERE tecdoc_id IN ({','.join(str(t) for t in chunk)})"),
        ).fetchall()
        for r in cat_rows:
            tecdoc_to_cat[r.tecdoc_id] = r.id

    all_pairs = set()
    for pairs in cat_map.values():
        for article, brand, _ in pairs:
            all_pairs.add((article, brand))

    parts_lookup = {}
    pair_list = list(all_pairs)
    for i in range(0, len(pair_list), 500):
        chunk = pair_list[i:i + 500]
        conditions = []
        params = {}
        for j, (article, brand) in enumerate(chunk):
            conditions.append(f"(article = :art_{j} AND COALESCE(brand, '') = :br_{j})")
            params[f"art_{j}"] = article
            params[f"br_{j}"] = brand
        p_rows = db.execute(
            text(f"SELECT id, article, COALESCE(brand, '') AS brand, category_id FROM parts WHERE {' OR '.join(conditions)}"),
            params,
        ).fetchall()
        for r in p_rows:
            parts_lookup[(r.article, r.brand)] = (r.id, r.category_id)

    updates = []
    for tecdoc_id, article_brands in cat_map.items():
        cat_id = tecdoc_to_cat.get(tecdoc_id)
        if cat_id is None:
            continue
        for article, brand, name in article_brands:
            part_info = parts_lookup.get((article, brand))
            if not part_info:
                continue
            # Don't overwrite manually set categories
            if part_info[1] is not None:
                continue
            # Keyword-based sub-categorization for belts
            resolved_cat_id = cat_id
            name_lower = name.lower()
            if "грм" in name_lower:
                if "комплект" in name_lower:
                    resolved_cat_id = 149  # Комплект ремня ГРМ
                else:
                    resolved_cat_id = 148  # Ремень ГРМ
            updates.append({"id": part_info[0], "category_id": resolved_cat_id})

    if not updates:
        db.commit()
        return 0

    # Batch update
    for i in range(0, len(updates), 500):
        chunk = updates[i:i + 500]
        db.execute(
            text("UPDATE parts SET category_id = :category_id WHERE id = :id"),
            chunk,
        )
    db.commit()
    return len(updates)
